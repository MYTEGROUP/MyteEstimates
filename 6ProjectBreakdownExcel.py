import json
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill, NamedStyle
from openpyxl.utils.dataframe import dataframe_to_rows

# Load JSON data
with open('storage/ProjectBreakdown1.json') as file:
    data = json.load(file)

# Flatten the data for Excel conversion
rows = []
for stakeholder, epics in data.items():
    for epic_id, stories in epics.items():
        for story in stories:
            acceptance_criteria = "\n".join(story['Story Acceptance Criteria'])  # Join all criteria
            for task in story['Tasks']:
                row = {
                    'Stakeholder': stakeholder,  # Include Stakeholder
                    'EpicID': epic_id,
                    'EPIC TITLE': story['Epic Title'],
                    'EPIC DESCRIPTION': story['Epic Description'],
                    'Story ID': story['Story ID'],
                    'Story Title': story['Story Title'],
                    'Story Description': story['Story Description'],
                    'Story Acceptance Criteria': acceptance_criteria,
                    'Task ID': task['Task ID'],
                    'Task Description': task['Description'],
                    'Estimated Hours': task['Estimated Hours'],  # Include Estimated Hours
                    'Cost': task['Cost']  # Include Cost
                }
                rows.append(row)

# Create DataFrame
df = pd.DataFrame(rows)

# Create a new Excel file with the DataFrame
wb = openpyxl.Workbook()
ws = wb.active

# Append DataFrame rows to Excel sheet
for r in dataframe_to_rows(df, index=False, header=True):
    ws.append(r)

# Custom Style for readability
header_style = NamedStyle(name="header_style", font=Font(bold=True, color="FFFFFF"), fill=PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid"))
ws.row_dimensions[1].height = 20  # Set header row height
for cell in ws[1]:
    cell.style = header_style

# Style and merge logic
def apply_merge_style(ws, column_indices):
    """
    Applies merging and formatting to specified columns based on content continuity.
    """
    for idx in column_indices:
        current_cell = ws.cell(row=2, column=idx)
        start_row = 2
        for row in range(3, ws.max_row + 1):
            cell = ws.cell(row=row, column=idx)
            if cell.value != current_cell.value or row == ws.max_row:
                end_row = row - 1 if cell.value != current_cell.value else row
                ws.merge_cells(start_row=start_row, start_column=idx, end_row=end_row, end_column=idx)
                for r in range(start_row, end_row + 1):
                    ws.cell(row=r, column=idx).alignment = Alignment(vertical='top', horizontal='left')
                current_cell = cell
                start_row = row

# Apply merging to all relevant columns including the new Stakeholder column
apply_merge_style(ws, [1, 2, 3, 4, 5, 6, 7, 8])  # Adjust the indices to include all the columns that need merging

# Save the styled Excel file
wb.save('StyledProjectBreakdown.xlsx')
